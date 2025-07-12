import openpyxl
import os
import json

def fill_excel_with_json_data(folder_path, excel_file):
    # Load Excel file
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    # Iterate over JSON files in the folder
    for filename in os.listdir(folder_path):
        print(filename)
        file_path = os.path.join(folder_path, filename)
        if filename.endswith('.json'):
            with open(file_path, 'r') as file:
                try:
                    data = json.load(file)
                    placement = data.get("placement", {})
                    latency_dict = data.get("latencyPerTupleType", {})
                    
                    # Update corresponding cells in Excel sheet based on placement dictionary
                    for row_name, col_name in placement.items():
                        if row_name == 'average':
                            latency = latency_dict['average_TUPLE']
                        elif row_name == 'blobRead':
                            latency = latency_dict['blobRead_TUPLE']
                        elif row_name == 'decisionTree':
                            latency = latency_dict['decisionTree_TUPLE']
                        elif row_name == 'errorEstimate':
                            latency = latency_dict['errorEstimate_TUPLE']
                        elif row_name == 'mqttPublish':
                            latency = latency_dict['mqttPublish_TUPLE']
                        elif row_name == 'source':
                            latency = latency_dict['TUPLE']
                        elif row_name == 'multiVarLinearReg':
                            latency = latency_dict['multiVarLinearReg_TUPLE']
                        elif row_name == 'senMLParse':
                            latency = latency_dict['senMLParse_TUPLE']
                        else:
                            latency = latency_dict['sink_TUPLE']
                        row_idx = find_row_index(sheet, row_name)
                        col_idx = find_column_index(sheet, col_name)
                        if row_idx is not None and col_idx is not None:
                            sheet.cell(row=row_idx, column=col_idx).value = latency
                except json.JSONDecodeError as e:
                    print(f"Error decoding JSON file '{filename}': {e}")
                except Exception as e:
                    print(f"Error processing JSON file '{filename}': {e}")

    wb.save(excel_file)

def find_row_index(sheet, row_name):
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=1).value == row_name:
            return row_idx
    return None

def find_column_index(sheet, col_name):
    for col_idx in range(2, sheet.max_column + 1):
        if sheet.cell(row=1, column=col_idx).value == col_name:
            return col_idx
    return None

# Folder containing JSON files
folder_path = "pred_sim_files_31/data"
# Excel file to be updated
excel_file = "pred_dataflow.xlsx"

fill_excel_with_json_data(folder_path, excel_file)
print(f"Excel file '{excel_file}' updated successfully.")
