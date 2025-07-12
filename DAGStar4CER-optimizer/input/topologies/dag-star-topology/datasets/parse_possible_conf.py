import json
import openpyxl

def extract_operator_names_from_json(workflow_json_file_path):
    operator_names = []
     # Open the JSON file
    with open(workflow_json_file_path, 'r') as file:
        # Load JSON data
        data = json.load(file)

        # Check if "sites" key exists in the data
        if "operators" in data:
            # Iterate over each site dictionary
            for operator in data["operators"]:
                # Check if "siteName" key exists in the site dictionary
                if "name" in operator:
                    # Append the value of "siteName" to the list
                    operator_names.append(operator["name"])
                else:
                    print("Warning: 'name' key not found in one of the oepartor dictionaries.")
        else:
            print("Error: 'operators' key not found in the JSON data.")

    return operator_names


def extract_site_names_from_json(network_json_file_path):
    site_names = []

    # Open the JSON file
    with open(network_json_file_path, 'r') as file:
        # Load JSON data
        data = json.load(file)

        # Check if "sites" key exists in the data
        if "sites" in data:
            # Iterate over each site dictionary
            for site in data["sites"]:
                # Check if "siteName" key exists in the site dictionary
                if "siteName" in site:
                    # Append the value of "siteName" to the list
                    site_names.append(site["siteName"])
                else:
                    print("Warning: 'siteName' key not found in one of the site dictionaries.")
        else:
            print("Error: 'sites' key not found in the JSON data.")

    return site_names

def create_excel(lst1, lst2, output_file):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Write headers
    for j, col_value in enumerate(lst2, start=2):
        sheet.cell(row=1, column=j).value = col_value

    # Write data
    for i, row_value in enumerate(lst1, start=2):
        sheet.cell(row=i, column=1).value = row_value
        for j, col_value in enumerate(lst2, start=2):
            sheet.cell(row=i, column=j).value = -1

    wb.save(output_file)

# Extract the possible locations, i.e., sites for this simulation from the appropriate JSON file
conf_num = 2047
network_json_file_path = f'/Users/dbanelas/Developer/CREXDATA/placement-simulation-suite/files/paper-experiments-simulations/exp_{conf_num}/network_{conf_num}_1_optimizer.json'  # Provide the path to your JSON file
workflow_json_file_path = "/Users/dbanelas/Developer/CREXDATA/placement-simulation-suite/files/paper-experiments-simulations/workflows/riot-etl-ifogsim.json"  # Provide the path to your JSON file

site_names = extract_site_names_from_json(network_json_file_path)
operator_names = extract_operator_names_from_json(workflow_json_file_path)

output_file = f'etl_xlsx/etl_{conf_num}_avg.xlsx'

create_excel(operator_names, site_names, output_file)