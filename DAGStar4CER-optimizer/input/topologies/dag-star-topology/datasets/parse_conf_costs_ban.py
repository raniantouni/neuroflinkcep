import openpyxl
import os
import json

def fill_excel_with_json_data(folder_path, excel_file, mode = "avg"):
    # Load Excel file
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    operator_config_latency_dict = {}

    # Iterate over JSON files in the folder
    for filename in os.listdir(folder_path):
        print(filename)
        file_path = os.path.join(folder_path, filename)
        if filename.endswith('.json'):
            with open(file_path, 'r') as file:
                try:
                    data = json.load(file)
                    placement = data.get("placement", {})

                    latency_dict = data.get("latencyPerTupleType", {})

                    for operator, config in placement.items():
                        operator_tuple = operator + "_TUPLE" if operator != "source" else "TUPLE"
                        operator_latency_in_config = latency_dict[operator_tuple]

                        if operator in operator_config_latency_dict:
                            if config in operator_config_latency_dict[operator]:
                                operator_config_latency_dict[operator][config].append(operator_latency_in_config)
                            else:
                                operator_config_latency_dict[operator][config] = []
                                operator_config_latency_dict[operator][config].append(operator_latency_in_config)
                        else:    
                            operator_config_latency_dict[operator] = {}
                            operator_config_latency_dict[operator][config] = []
                            operator_config_latency_dict[operator][config].append(operator_latency_in_config)
                    
                except json.JSONDecodeError as e:
                    print(f"Error decoding JSON file '{filename}': {e}")
                except Exception as e:
                    print(f"Error processing JSON file '{filename}': {e}")

    for operator in operator_config_latency_dict.keys():
        latency_per_config = operator_config_latency_dict[operator]
        row_idx = find_row_index(sheet, operator)      
        for config in latency_per_config:
            col_idx = find_column_index(sheet, config)
            latency_list = operator_config_latency_dict[operator][config]
            # latency = min(latency_list)
            if mode == "avg":
                latency = sum(latency_list) / len(latency_list)
            elif mode == "min":
                latency = min(latency_list)
            elif mode == "max":
                latency = max(latency_list)
            
            if row_idx is not None and col_idx is not None:
                sheet.cell(row=row_idx, column=col_idx).value = latency
            

    wb.save(excel_file)

def find_row_index(sheet, row_name):
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=1).value == row_name:
            return row_idx
    return None

def find_column_index(sheet, col_name):
    for col_idx in range(2, sheet.max_column + 1):
        if sheet.cell(row=1, column=col_idx).value == col_name:
            return col_idx
    return None

conf_num = 2047
# Folder containing JSON files
folder_path = f'stats/{conf_num}'
# Excel file to be updated
mode = "avg"
excel_file = f'stats_xlsx/stats_{conf_num}_{mode}.xlsx'

fill_excel_with_json_data(folder_path, excel_file, mode)
print(f"Excel file '{excel_file}' updated successfully.")
